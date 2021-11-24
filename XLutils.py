
import openpyxl

def numrows(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.max_row

def nuumcol(file,sheetname):
    wokbook=openpyxl.load_workbook(file)
    sheet=wokbook.get_sheet_by_name(sheetname)
    return sheet.max_column

def getdata(file,sheetname,row,col):
    wokbook = openpyxl.load_workbook(file)
    sheet = wokbook.get_sheet_by_name(sheetname)
    return sheet.cell(row,col).value

def editdata(file,sheetname,row,col,valu):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    sheet.cell(row,col).value=valu
    workbook.save(file)

