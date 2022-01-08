
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
import re

options = webdriver.ChromeOptions()
prefs = {"download.default_directory" : "F:\pysel\data"}
options.add_experimental_option("prefs", prefs)

workbook=openpyxl.load_workbook("F:/pycharm practice/nopcommerceApp/database/Book2.xlsx")
sheet=workbook.active
No_rows=sheet.max_row
No_cols=sheet.max_column

def fda_files(value):

    a=value
    res_str = a.replace('"', '')
    driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=options)

    driver.get("https://www.accessdata.fda.gov/scripts/cder/daf/index.cfm")
    driver.implicitly_wait(10)
    driver.find_element_by_xpath("//*[@id='searchterm']").send_keys(res_str)
    driver.find_element_by_xpath("//*[@id='DrugNameform']/div[2]/button[1]").click()

    tables=driver.find_elements_by_xpath("//*[@id='accordion']/div/div/h4/a")

    for table in tables:
        driver.find_element_by_link_text(table.text).click()
    files=driver.find_elements_by_xpath("//span[text()='CSV']")

    for file in files:
        driver.find_element_by_link_text("CSV").click()


    time.sleep(10)
    driver.quit()


for r in range(2,No_rows+1):
    value=sheet.cell(r,1).value
    fda_files(value)
